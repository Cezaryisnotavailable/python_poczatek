import pandas as pd
from openpyxl import load_workbook

location = ('C:\\Users\\48697\Documents\\tickers.xlsx')

workbook = load_workbook(filename=location)
first_sheet = workbook.active

row = first_sheet.max_row + 1



def tickers_list():
    tickers_list = []
    for i in range(1, row):
        tickers_list.append(first_sheet[("A" + str(i))].value)
    return tickers_list



