import requests
from tickers import tickers_list
from openpyxl import load_workbook

lista = tickers_list()
print(lista)

location = ('C:\\Users\\48697\Documents\\final_tickers.xlsx')
workbook = load_workbook(filename=location)
sheet_1 = workbook['Freefloat']



def run_example():



    LICZBA_SPACJI = 73

    row_number = 1

    for ticker in lista:
        url = "https://www.bankier.pl/gielda/notowania/new-connect/" + ticker + "/akcjonariat"
        try:
            response = requests.get(url)
        except requests.RequestException as error:
            print(f"Błąd przy połączeniu: {error}")
            return

        try:
            response.raise_for_status()
        except requests.HTTPError as error:
            print(f"Żądanie zakończone niepowodzeniem {error}")
            return


        text = response.text
        razem_index = text.find("Free float:")
        index_start_from = razem_index + LICZBA_SPACJI
        index_stop = index_start_from + 5
        wanted_text = text[index_start_from:index_stop]
        # print(ticker, wanted_text, url)
        # freefloat_list.append([ticker, wanted_text])

        sheet_1.cell(row=row_number, column=1, value=ticker)
        sheet_1.cell(row=row_number, column=2, value=wanted_text)
        sheet_1.cell(row=row_number, column=3, value=url)


        row_number += 1
    workbook.save('C:\\Users\\48697\Documents\\final_tickers.xlsx')









if __name__ == "__main__":

    print(run_example())